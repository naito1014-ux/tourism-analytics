#!/usr/bin/env python3
"""
宿泊旅行統計ダッシュボード データ更新スクリプト

使い方:
  python3 update_data.py --shukuhaku 001998942.xlsx --jnto 20260415_1615-5.xlsx --consumption 001992581__1_.xls --natpref 001905499.xlsx

各ファイルは省略可能。指定されたファイルのみ更新し、既存の data.json とマージします。
"""
import argparse, json, os, sys

def parse_shukuhaku(filepath):
    """観光庁 宿泊旅行統計 推移表"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    
    jp_year_map = {
        '平成23年':2011,'平成24年':2012,'平成25年':2013,'平成26年':2014,
        '平成27年':2015,'平成28年':2016,'平成29年':2017,'平成30年':2018,
        '令和元年':2019,'令和2年':2020,'令和3年':2021,'令和4年':2022,
        '令和5年':2023,'令和6年':2024,'令和7年':2025
    }
    jp_months = {'1月':1,'2月':2,'3月':3,'4月':4,'5月':5,'6月':6,'7月':7,'8月':8,'9月':9,'10月':10,'11月':11,'12月':12}
    
    def parse_old(sheet_name):
        ws = wb[sheet_name]; rows = list(ws.iter_rows(values_only=True))
        yr = list(rows[2]); mr = list(rows[3])
        cy = None; cols = []
        for i in range(1, len(yr)):
            if yr[i]: cy = jp_year_map.get(str(yr[i]).strip())
            ms = mr[i] if i < len(mr) else None
            if ms and cy:
                try: cols.append((cy, int(str(ms).replace('月','')), i))
                except: pass
        data = {}
        for row in rows[4:]:
            vals = list(row); p = vals[0]
            if not p: continue
            p = str(p).strip()
            if p.startswith('※'): continue
            pd = {}
            for y,m,ci in cols:
                if y < 2023: continue
                v = vals[ci] if ci < len(vals) else None
                if v is not None and v != '-':
                    try: pd[f"{y}{m:02d}"] = round(float(v))
                    except: pass
            data[p] = pd
        return data
    
    def parse_new(sheet_name):
        ws = wb[sheet_name]; rows = list(ws.iter_rows(values_only=True))
        mr = list(rows[3]); cols = []
        for i in range(1, len(mr)):
            ms = mr[i]
            if ms and str(ms).strip() in jp_months:
                cols.append((2026, jp_months[str(ms).strip()], i))
        data = {}
        for row in rows[4:]:
            vals = list(row); p = vals[0]
            if not p: continue
            p = str(p).strip()
            if p.startswith('※'): continue
            pd = {}
            for y,m,ci in cols:
                v = vals[ci] if ci < len(vals) else None
                if v is not None and v != '-':
                    try: pd[f"{y}{m:02d}"] = round(float(v))
                    except: pass
            data[p] = pd
        return data
    
    def merge(old, new):
        m = {}
        for k in set(list(old.keys())+list(new.keys())):
            m[k] = {**old.get(k,{}), **new.get(k,{})}
        return m
    
    def parse_occ_old(sheet_name):
        ws = wb[sheet_name]; rows = list(ws.iter_rows(values_only=True))
        yr = list(rows[2]); mr = list(rows[3])
        cy = None; cols = []
        for i in range(2, len(yr)):
            if yr[i]: cy = jp_year_map.get(str(yr[i]).strip())
            ms = mr[i] if i < len(mr) else None
            if ms and cy:
                try: cols.append((cy, int(str(ms).replace('月','')), i))
                except: pass
        data = {}; cp = None
        for row in rows[4:]:
            vals = list(row)
            if vals[0]: cp = str(vals[0]).strip()
            if not cp or cp.startswith('※'): continue
            ft = vals[1]
            if ft and str(ft).strip().replace('\n','') == '計':
                pd = {}
                for y,m,ci in cols:
                    if y < 2023: continue
                    v = vals[ci] if ci < len(vals) else None
                    if v is not None and v != '-':
                        try: pd[f"{y}{m:02d}"] = round(float(v), 1)
                        except: pass
                data[cp] = pd
        return data

    def parse_occ_new(sheet_name):
        ws = wb[sheet_name]; rows = list(ws.iter_rows(values_only=True))
        mr = list(rows[3]); cols = []
        for i in range(2, len(mr)):
            ms = mr[i]
            if ms and str(ms).strip() in jp_months:
                cols.append((2026, jp_months[str(ms).strip()], i))
        data = {}; cp = None
        for row in rows[4:]:
            vals = list(row)
            if vals[0]: cp = str(vals[0]).strip()
            if not cp or cp.startswith('※'): continue
            ft = vals[1]
            if ft and '計' in str(ft).strip().replace('\n',''):
                pd = {}
                for y,m,ci in cols:
                    v = vals[ci] if ci < len(vals) else None
                    if v is not None and v != '-':
                        try: pd[f"{y}{m:02d}"] = round(float(v), 1)
                        except: pass
                data[cp] = pd
        return data
    
    t = merge(parse_old('旧1-2'), parse_new('1-1'))
    j = merge(parse_old('旧2-2'), parse_new('2-1'))
    f = merge(parse_old('旧3-2'), parse_new('3-1'))
    o = merge(parse_occ_old('旧4-2'), parse_occ_new('4-1'))
    
    prefs = sorted([k for k in t if k != '全　国'], key=lambda x: x[:2])
    return {"t":t, "j":j, "f":f, "o":o, "p":["全　国"]+prefs}


def parse_jnto(filepath):
    """JNTO 訪日外客数"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    KEY = ['韓国','中国','台湾','香港','タイ','シンガポール','マレーシア','インドネシア',
           'フィリピン','ベトナム','インド','米国','カナダ','豪州','英国','フランス','ドイツ',
           'イタリア','スペイン','メキシコ','ニュージーランド','ロシア','総数']
    d, g = {}, {}
    for sy in ['2023','2024','2025','2026']:
        if sy not in wb.sheetnames: continue
        ws = wb[sy]; rows = list(ws.iter_rows(values_only=True)); y = int(sy)
        for row in rows[4:]:
            vals = list(row)
            c = str(vals[0]).strip() if vals[0] else (str(vals[1]).strip().lstrip('　 ') if vals[1] else None)
            if not c or c.startswith('注') or c.startswith('＊'): continue
            if c not in KEY: continue
            if c not in d: d[c], g[c] = {}, {}
            for m in range(1, 13):
                cc, gc = 2+(m-1)*2, 3+(m-1)*2
                if cc < len(vals) and vals[cc] is not None:
                    try: d[c][f"{y}{m:02d}"] = round(float(vals[cc]))
                    except: pass
                if gc < len(vals) and vals[gc] is not None:
                    try: g[c][f"{y}{m:02d}"] = round(float(vals[gc]), 1)
                    except: pass
    return {"d":d, "g":g, "countries":[c for c in KEY if c in d], "regions":[]}


def parse_jnto_monthly(filepath, whitelist):
    """JNTO 訪日外客数（月次様式 / YYYY.MM シート）。
    1ファイル＝1か月。国名=col2、当月値(2026年)=col5、前年同月(2025年)=col4。
    伸率は col6 が数式のため自前計算 (col5-col4)/col4*100。
    whitelist（＝data.json の jc 23件）に含まれる国・地域のみ取り込む（ホワイトリスト方式）。
    行が存在しない国（例: ニュージーランド）は自然に欠測となりスキップ（歯抜けを許容）。
    戻り値: {"d":{国:{YYYYMM:実数}}, "g":{国:{YYYYMM:伸率}}, "ym":"YYYYMM"}"""
    import re
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)

    # --- シート名 YYYY.MM を検出 ---
    ym = None; sheet = None
    for s in wb.sheetnames:
        m = re.fullmatch(r'(\d{4})\.(\d{1,2})', str(s).strip())
        if m:
            sheet = s
            ym = f"{int(m.group(1))}{int(m.group(2)):02d}"
            break
    if not sheet:
        raise ValueError(f"YYYY.MM 形式のシートが見つかりません: {wb.sheetnames}")

    ws = wb[sheet]; rows = list(ws.iter_rows(values_only=True))
    wl = set(whitelist)
    d, g = {}, {}
    for row in rows[6:]:                       # 総数行(r6)以降が国データ
        vals = list(row)
        c = vals[2] if len(vals) > 2 else None
        if c is None: continue
        c = str(c).strip()
        # 注記行ガード: 空 / 12字超 / ◆・注 始まり を除外
        if not c or len(c) > 12 or c.startswith('◆') or c.startswith('注'): continue
        if c not in wl: continue               # jc の23件のみ（北欧地域・中東地域・その他は無視）
        cur = vals[5] if len(vals) > 5 else None    # 当月 2026年
        prev = vals[4] if len(vals) > 4 else None   # 前年同月 2025年
        if cur is None: continue               # 値欠測はスキップ
        try:
            v = round(float(cur))
        except (TypeError, ValueError):
            continue
        d.setdefault(c, {})[ym] = v            # jd（実数）差分マージ用
        # 伸率 jg を自前計算（前年同月が有効な数値のときのみ）
        try:
            pv = float(prev)
            if pv != 0:
                g.setdefault(c, {})[ym] = round((v - pv) / pv * 100, 1)
        except (TypeError, ValueError):
            pass
    return {"d":d, "g":g, "ym":ym}


def parse_shukuhaku_kakuho(filepath):
    """観光庁 宿泊旅行統計 月次確報（第2次確報）。1ファイル＝1か月。
    第2表: col1=延べ宿泊者数(t)、col19=うち外国人延べ(f)、日本人(j)=t−f。
    第8表: col1=客室稼働率(o)。
    全国行(r6, col0='令和N年M月')は '全　国' キー、都道府県行(r7〜)は先頭の
    全角空白を除去してキー化（例 '　01北海道'→'01北海道'）。運輸局等の行は除外。
    案A: 全国 t/j/f/o も確報値で返す（呼び出し側で既存を上書き）。
    戻り値: {"t":{key:{ym:v}}, "j":..., "f":..., "o":..., "ym":ym}"""
    import re
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)

    def find_sheet(num):
        for s in wb.sheetnames:
            m = re.fullmatch(rf'第{num}表\((\d+)月\)', str(s).strip())
            if m: return s, int(m.group(1))
        return None, None

    s2, mon = find_sheet(2)
    s8, _ = find_sheet(8)
    if not s2 or not s8:
        raise ValueError(f"第2表/第8表が見つかりません: {wb.sheetnames}")

    r2 = list(wb[s2].iter_rows(values_only=True))
    r8 = list(wb[s8].iter_rows(values_only=True))

    # 年は全国行(r6) col0 の '令和N年M月' から取得（令和N → 2018+N）
    m = re.match(r'令和(\d+)年', str(r2[6][0]))
    if not m or mon is None:
        raise ValueError(f"年月を解析できません: {r2[6][0]!r}")
    ym = f"{int(m.group(1)) + 2018}{mon:02d}"

    def num(v):
        if v is None or v == '-': return None
        try: return float(v)
        except (TypeError, ValueError): return None

    def keyname(cell, is_total):
        if is_total: return '全　国'
        name = str(cell).replace('　', '').strip()
        return name if name[:2].isdigit() else None   # 47都道府県のみ（運輸局等を除外）

    t, j, f, o = {}, {}, {}, {}

    # 第2表 → t / f / j（j=t−f）
    for i in range(6, len(r2)):
        cell = r2[i][0]
        if not cell: continue
        key = keyname(cell, i == 6)
        if not key: continue
        tv = num(r2[i][1]); fv = num(r2[i][19])
        if tv is not None: t[key] = {ym: round(tv)}
        if fv is not None: f[key] = {ym: round(fv)}
        if tv is not None and fv is not None:
            j[key] = {ym: round(tv - fv)}

    # 第8表 → o（客室稼働率）
    for i in range(6, len(r8)):
        cell = r8[i][0]
        if not cell: continue
        key = keyname(cell, i == 6)
        if not key: continue
        ov = num(r8[i][1])
        if ov is not None: o[key] = {ym: round(ov, 1)}

    return {"t":t, "j":j, "f":f, "o":o, "ym":ym}


def parse_consumption(filepath):
    """インバウンド消費動向調査"""
    import pandas as pd
    xls = pd.ExcelFile(filepath, engine="xlrd")
    
    def get_cols(df, row=4):
        return [(str(df.iloc[row,c]).strip(), c) for c in range(4, df.shape[1], 2) if pd.notna(df.iloc[row,c])]
    
    df31 = pd.read_excel(xls, sheet_name='表3-1', header=None)
    ct = {}
    for n,c in get_cols(df31):
        v = df31.iloc[6, c+1]
        if pd.notna(v) and v != '-':
            try: ct[n] = round(float(v))
            except: pass
    
    df41 = pd.read_excel(xls, sheet_name='表4-1', header=None)
    cn = {}
    for n,c in get_cols(df41):
        v = df41.iloc[6, c+1]
        if pd.notna(v) and v != '-':
            try: cn[n] = round(float(v), 1)
            except: pass
    
    df21 = pd.read_excel(xls, sheet_name='表2-1', header=None)
    er = {'宿泊費':57,'飲食費':58,'交通費':59,'娯楽等サービス費':68,'買物代':81}
    ce = {}
    for n,c in get_cols(df21):
        items = {}
        for cat,r in er.items():
            v = df21.iloc[r, c+1]
            if pd.notna(v) and v != '-':
                try: items[cat] = round(float(v))
                except: pass
        ce[n] = items
    
    df33 = pd.read_excel(xls, sheet_name='表3-3', header=None)
    pt = {}
    for n,c in get_cols(df33):
        v = df33.iloc[6, c+1]
        if pd.notna(v) and v != '-':
            try: pt[n] = round(float(v))
            except: pass
    
    df34 = pd.read_excel(xls, sheet_name='表3-4', header=None)
    pn = {}
    for n,c in get_cols(df34):
        v = df34.iloc[6, c+1]
        if pd.notna(v) and v != '-':
            try: pn[n] = round(float(v))
            except: pass
    
    df43 = pd.read_excel(xls, sheet_name='表4-3', header=None)
    pg = {}
    for n,c in get_cols(df43):
        v = df43.iloc[6, c+1]
        if pd.notna(v) and v != '-':
            try: pg[n] = round(float(v), 1)
            except: pass
    
    df23 = pd.read_excel(xls, sheet_name='表2-3', header=None)
    per = {'宿泊費':24,'飲食費':25,'交通費':26,'娯楽等サービス費':30,'買物代':31}
    pe = {}
    for n,c in get_cols(df23):
        items = {}
        for cat,r in per.items():
            v = df23.iloc[r, c+1]
            if pd.notna(v) and v != '-':
                try: items[cat] = round(float(v))
                except: pass
        if items: pe[n] = items
    
    df61 = pd.read_excel(xls, sheet_name='表6-1', header=None)
    pr = []
    for i in range(7, min(60, df61.shape[0])):
        c2 = str(df61.iloc[i,2]) if pd.notna(df61.iloc[i,2]) else ""
        if c2.strip() and any(x in c2 for x in ['県','都','府','道']): pr.append((i, c2.strip()))
    pv = {}
    for cn2,col in get_cols(df61):
        rates = {}
        for r,pn2 in pr:
            v = df61.iloc[r, col+1]
            if pd.notna(v) and v != '-':
                try: rates[pn2] = round(float(v), 2)
                except: pass
        pv[cn2] = rates
    
    # Reverse: pref -> {country: rate}
    pv2 = {}
    for cn2, rates in pv.items():
        for pn2, rate in rates.items():
            if pn2 not in pv2: pv2[pn2] = {}
            pv2[pn2][cn2] = rate
    
    return {"year":2025,"ct":ct,"cn":cn,"ce":ce,"pt":pt,"pn":pn,"pg":pg,"pe":pe,"vr":pv,"pv":pv2}


def parse_natpref(filepath):
    """宿泊旅行統計 国籍別×都道府県"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    
    def parse_sheet(sheet_name):
        ws = wb[sheet_name]; rows = list(ws.iter_rows(values_only=True))
        nat_cols = []
        for c in range(2, len(list(rows[4]))):
            v = rows[4][c]
            if v and str(v).strip(): nat_cols.append((str(v).strip(), c))
        data = {}
        for i in range(6, min(70, len(rows))):
            p = str(rows[i][0]).strip().replace('\u3000','') if rows[i][0] else ""
            if not p or p.startswith('※') or '運輸局' in p or '管内計' in p: continue
            pd = {}
            total = rows[i][1]
            if total:
                ts = str(total).replace('*','').replace(',','').strip()
                try: pd['total'] = int(float(ts))
                except: pass
            for nn,c in nat_cols:
                v = rows[i][c]
                if v and v != '-':
                    vs = str(v).replace('*','').replace(',','').strip()
                    try: pd[nn] = int(float(vs))
                    except: pass
            if pd: data[p] = pd
        return data
    
    annual = parse_sheet('参考第1表(年計)')
    monthly = {}
    for m in range(1, 13):
        s = f'参考第1表({m}月)'
        if s in wb.sheetnames: monthly[m] = parse_sheet(s)
    
    result = {}
    all_p = set(list(annual.keys()))
    for m in monthly: all_p.update(monthly[m].keys())
    for p in sorted(all_p):
        if not (len(p)>=2 and p[:2].isdigit()): continue
        pe = {}
        for nat, val in annual.get(p, {}).items():
            pe[nat] = {"a": val}
        for m in range(1, 13):
            for nat, val in monthly.get(m, {}).get(p, {}).items():
                if nat not in pe: pe[nat] = {}
                pe[nat][str(m)] = val
        result[p] = pe
    return result


def main():
    parser = argparse.ArgumentParser(description='宿泊統計ダッシュボード データ更新')
    parser.add_argument('--shukuhaku', help='観光庁 宿泊旅行統計 推移表 (.xlsx)')
    parser.add_argument('--jnto', help='JNTO 訪日外客数 (.xlsx)')
    parser.add_argument('--consumption', help='インバウンド消費動向調査 (.xls)')
    parser.add_argument('--natpref', help='宿泊旅行統計 国籍別×都道府県 (.xlsx)')
    parser.add_argument('--output', default='data.json', help='出力先 (default: data.json)')
    args = parser.parse_args()
    
    # Load existing data if available
    existing = {}
    if os.path.exists(args.output):
        with open(args.output) as f:
            existing = json.load(f)
        print(f"既存の {args.output} を読み込みました")
    
    if args.shukuhaku:
        print(f"宿泊旅行統計を処理中: {args.shukuhaku}")
        s = parse_shukuhaku(args.shukuhaku)
        existing.update({"t":s["t"],"j":s["j"],"f":s["f"],"o":s["o"],"p":s["p"]})
        print(f"  → {len(s['p'])} 都道府県")
    
    if args.jnto:
        print(f"JNTO訪日外客数を処理中: {args.jnto}")
        j = parse_jnto(args.jnto)
        existing.update({"jd":j["d"],"jg":j["g"],"jc":j["countries"],"jr":j["regions"]})
        print(f"  → {len(j['countries'])} カ国")
    
    if args.consumption:
        print(f"消費動向調査を処理中: {args.consumption}")
        c = parse_consumption(args.consumption)
        existing["cs"] = c
        print(f"  → {len(c['ct'])} カ国, {len(c['pt'])} 都道府県")
    
    if args.natpref:
        print(f"国籍別×都道府県を処理中: {args.natpref}")
        n = parse_natpref(args.natpref)
        existing["np"] = n
        print(f"  → {len(n)} 都道府県")
    
    with open(args.output, 'w') as f:
        json.dump(existing, f, ensure_ascii=False, separators=(',',':'))
    
    size = os.path.getsize(args.output)
    print(f"\n✅ {args.output} を更新しました ({size/1024:.0f} KB)")

if __name__ == '__main__':
    main()
