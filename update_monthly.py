#!/usr/bin/env python3
"""
月次速報データでダッシュボードを更新するスクリプト

使い方:
  python3 update_monthly.py 001992573__2_.xlsx

観光庁の月次速報Excelを指定すると:
1. data.json を更新（全国データを差分追加）
2. index.html を再生成（更新データを埋め込み）
3. git add / commit / push（--deploy オプション時）
"""
import json, re, sys, os, subprocess

def main():
    if len(sys.argv) < 2:
        print("使い方: python3 update_monthly.py <月次速報.xlsx> [--deploy]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    auto_deploy = "--deploy" in sys.argv
    
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    
    # Auto-detect month
    file_month = None
    for s in wb.sheetnames:
        m = re.search(r'第1表\((\d+)月\)', s)
        if m:
            file_month = int(m.group(1))
            break
    
    if not file_month:
        print("❌ 月次速報のフォーマットが認識できません")
        sys.exit(1)
    
    # Parse date
    ws1 = wb[f'第1表({file_month}月)']
    rows1 = list(ws1.iter_rows(values_only=True))
    date_str = str(rows1[6][0]).strip()
    m = re.match(r'(令和)(\d+)年(\d+)月', date_str)
    if not m:
        print(f"❌ 日付を解析できません: {date_str}")
        sys.exit(1)
    
    year = int(m.group(2)) + 2018
    month = int(m.group(3))
    ym = f"{year}{month:02d}"
    
    # Extract data
    total = int(float(rows1[6][1]))
    foreign = int(float(rows1[6][2]))
    japanese = total - foreign
    
    ws5 = wb[f'第5表({file_month}月)']
    rows5 = list(ws5.iter_rows(values_only=True))
    occ = round(float(rows5[6][1]), 1)
    
    print(f"📊 {year}年{month}月データを検出")
    print(f"  延べ宿泊者数: {total:,}")
    print(f"  日本人: {japanese:,}")
    print(f"  外国人: {foreign:,}")
    print(f"  稼働率: {occ}%")
    
    # Load data.json
    data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
    with open(data_path) as f:
        data = json.load(f)
    
    # Update national data
    data["t"]["全　国"][ym] = total
    data["j"]["全　国"][ym] = japanese
    data["f"]["全　国"][ym] = foreign
    data["o"]["全　国"][ym] = occ
    
    # Save data.json
    with open(data_path, 'w') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',',':'))
    print(f"\n✅ data.json 更新完了")
    
    # Rebuild index.html
    index_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html')
    with open(index_path) as f:
        old_html = f.read()
    
    # Replace the data in index.html
    data_json_str = json.dumps(data, ensure_ascii=False, separators=(',',':'))
    
    # Find and replace the RAW data block
    raw_start = old_html.index('var RAW = ') + len('var RAW = ')
    raw_end = old_html.index(';\n</script>', raw_start)
    new_html = old_html[:raw_start] + data_json_str + old_html[raw_end:]
    
    with open(index_path, 'w') as f:
        f.write(new_html)
    print(f"✅ index.html 再生成完了 ({os.path.getsize(index_path)/1024:.0f} KB)")
    
    # Auto deploy
    if auto_deploy:
        print(f"\n🚀 デプロイ中...")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        subprocess.run(["git", "add", "data.json", "index.html"], cwd=script_dir)
        subprocess.run(["git", "commit", "-m", f"{year}年{month}月データ更新"], cwd=script_dir)
        subprocess.run(["git", "push"], cwd=script_dir)
        print(f"✅ デプロイ完了！")
    else:
        print(f"\nデプロイするには:")
        print(f'  git add data.json index.html')
        print(f'  git commit -m "{year}年{month}月データ更新"')
        print(f'  git push')

if __name__ == '__main__':
    main()
